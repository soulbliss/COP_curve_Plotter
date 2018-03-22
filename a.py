from __future__ import print_function
import wx
import gui
import openpyxl
import matplotlib.pyplot as plt
import numpy as np
from tkinter import *
from tkinter import messagebox



class Cursor(object):
    def __init__(self, ax):
        self.ax = ax
        self.lx = ax.axhline(color='k')  # the horiz line
        self.ly = ax.axvline(color='k')  # the vert line

        # text location in axes coords
        self.txt = ax.text(0.7, 0.9, '', transform=ax.transAxes)

    def mouse_move(self, event):
        if not event.inaxes:
            return

        x, y = event.xdata, event.ydata
        # update the line positions
        self.lx.set_ydata(y)
        self.ly.set_xdata(x)

        self.txt.set_text('x=%1.2f, y=%1.2f' % (x, y))
        plt.draw()


class SnaptoCursor(object):
    def __init__(self, ax, x, y):
        self.ax = ax
        self.lx = ax.axhline(color='b')  # the horiz line
        self.ly = ax.axvline(color='b')  # the vert line
        self.x = x
        self.y = y
        # text location in axes coords
        self.txt = ax.text(0.7, 0.9, '', transform=ax.transAxes)

    def mouse_move(self, event):

        if not event.inaxes:
            return

        x, y = event.xdata, event.ydata

        indx = np.searchsorted(self.x, [x])[0]
        x = self.x[indx]
        y = self.y[indx]
        # update the line positions
        self.lx.set_ydata(y)
        self.ly.set_xdata(x)

        self.txt.set_text('x=%1.2f, y=%1.2f' % (x, y))
        print('x=%1.2f, y=%1.2f' % (x, y))
        plt.draw()


















wb = openpyxl.load_workbook('R22.xlsx')
sheet = wb.get_sheet_by_name('data')

# to get the row number of input temperature

temp_row = checker = 0




class ethtFrame(gui.Heady):



	#constructor
	def __init__(self,parent):
		gui.Heady.__init__(self,parent)

	#initialize parent class
	

	def To_accep_temp(self,event):
		global h_all,p_all,h1,h2,h3,h4,p1
		global p2,cop,t,temp_list
	
		evap_temp =  eval(self.tect.GetValue())
		self.tect.SetValue (str(evap_temp))
		for x in range(1, 43, 1):
			checker = sheet.cell(row=x, column=1).value
			if checker == evap_temp:
				temp_row = x
				nirvana = x
				break
		h1 = sheet.cell(row=temp_row, column=3).value
		h2 = sheet.cell(row=temp_row, column=5).value
		h3 = sheet.cell(row=temp_row, column=6).value
		h4 = sheet.cell(row=temp_row, column=7).value
		p1 = sheet.cell(row=temp_row, column=2).value
		p2 = 15.00
		cop = sheet.cell(row=temp_row, column=9).value
		t = sheet.cell(row=temp_row, column=1).value
		h_all = [h1, h2, h3, h4]
				
		p_all = [p1, p2, p2, p1]
		
		temp_list = []
		cop = []

		cop_value = 0
		for x in range(-25, 15, 1):
			temp_list.append(x)
		for x in range(2, 42, 1):
			cop_value = sheet.cell(row=x, column=9).value
			cop.append(cop_value)


	def pvsh(self,event):
		fig = plt.figure()
		ax = fig.add_subplot(111)
		plt.plot(h_all, p_all)
		for xy in zip(h_all, p_all):
			ax.annotate('(%.7s, %.7s)' % xy, xy=xy, textcoords='data')
		plt.plot([h1, h4], [p1, p1])
		plt.axis([0, 400, 0, 20])
		plt.title('Pressure vs Enthalpy Graph')
		plt.ylabel('P - Pressure ( bars )')
		plt.xlabel('H - enthalpy ( kJ/kg)')
		plt.show()

	def copvstemp( self, event ):
		fig, ax = plt.subplots()
		cursor = SnaptoCursor(ax, temp_list, cop)
		plt.connect('motion_notify_event', cursor.mouse_move)
		ax.plot(temp_list, cop, linewidth=2.0)
		plt.axis([-30, 20, 0, 15])
		plt.title('COP vs Temp Graph')
		plt.ylabel('COP')
		plt.xlabel('Temp - °C')
		plt.show()



app = wx.App(False)

frame = ethtFrame(None)
#show the frame
frame.Show(True)
#start the applications
app.MainLoop()
